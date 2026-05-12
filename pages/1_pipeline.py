"""
Pipeline Page — Mobile.de Automation
Handles file uploads, runs the 3-step pipeline, returns downloadable Template.
"""
import io
import os
import shutil
import sys
import tempfile
from datetime import datetime
from pathlib import Path

import pandas as pd
import streamlit as st

# ── Page config ──────────────────────────────────────────────────────────────
st.set_page_config(page_title="Run Pipeline · Mobile.de", page_icon="⚙️", layout="wide")

st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;600&family=IBM+Plex+Sans:wght@300;400;600;700&display=swap');
    html, body, [class*="css"] { font-family: 'IBM Plex Sans', sans-serif; }

    .page-title { font-family:'IBM Plex Mono',monospace; font-size:2rem; font-weight:600; color:#fff; margin-bottom:.2rem; }
    .page-sub   { font-size:.9rem; color:#666; margin-bottom:2rem; }
    .accent     { color:#e8ff4d; }

    .upload-zone {
        background:#161616; border:1px dashed #333; border-radius:8px;
        padding:1.2rem 1.4rem; margin-bottom:.8rem;
    }
    .upload-label { font-size:.72rem; color:#e8ff4d; font-family:'IBM Plex Mono',monospace;
        letter-spacing:.15em; text-transform:uppercase; margin-bottom:.4rem; }
    .upload-hint  { font-size:.8rem; color:#555; margin-top:.3rem; }

    .log-box { background:#0a0a0a; border:1px solid #1e1e1e; border-radius:6px;
        padding:1rem 1.2rem; font-family:'IBM Plex Mono',monospace; font-size:.78rem;
        color:#aaa; line-height:1.7; max-height:320px; overflow-y:auto; }
    .log-ok   { color:#e8ff4d; }
    .log-err  { color:#ff5555; }
    .log-info { color:#888; }

    .stButton>button { background:#e8ff4d; color:#0d0d0d; border:none;
        font-family:'IBM Plex Mono',monospace; font-weight:600; font-size:.85rem;
        letter-spacing:.05em; padding:.6rem 1.4rem; border-radius:4px; }
    .stButton>button:hover { background:#fff; }

    .result-card { background:#0f1a00; border:1px solid #e8ff4d44; border-radius:8px;
        padding:1.4rem; margin-top:1.2rem; }
    .result-title { font-family:'IBM Plex Mono',monospace; font-size:.8rem; color:#e8ff4d;
        letter-spacing:.15em; text-transform:uppercase; margin-bottom:.5rem; }

    [data-testid="stSidebar"] { background:#111; border-right:1px solid #2a2a2a; }
</style>
""", unsafe_allow_html=True)

# ── Helpers ───────────────────────────────────────────────────────────────────

def save_upload(upload, dest: Path) -> None:
    dest.write_bytes(upload.read())


def run_pipeline(
    date_str: str,
    export_files,
    template_file,
    crawl_file=None,
) -> tuple[bytes | None, list[str]]:
    """
    Runs the full pipeline inside a temp directory.
    Returns (output_bytes, log_lines).
    """
    logs: list[str] = []

    with tempfile.TemporaryDirectory() as tmpdir:
        base = Path(tmpdir)

        # Copy scripts next to the temp workspace so BASE_DIR resolution works
        scripts_src = Path(__file__).resolve().parent.parent / "scripts"
        for sc in scripts_src.glob("*.py"):
            shutil.copy(sc, base / sc.name)

        # Build folder structure: <tmpdir>/<date_str>/exports/
        exports_dir = base / date_str / "exports"
        exports_dir.mkdir(parents=True, exist_ok=True)
        deliver_dir = base / date_str / "deliver"
        deliver_dir.mkdir(parents=True, exist_ok=True)

        # Save Template_file.xlsx at base level (scripts look for it there)
        tpl_dest = base / "Template_file.xlsx"
        save_upload(template_file, tpl_dest)
        logs.append(f'<span class="log-ok">✓</span> Template file saved → Template_file.xlsx')

        # Save export files
        for uf in export_files:
            dest = exports_dir / uf.name
            uf.seek(0)
            dest.write_bytes(uf.read())
            logs.append(f'<span class="log-ok">✓</span> Export saved → {uf.name}')

        # Save crawl file if provided
        crawl_path: Path | None = None
        if crawl_file:
            crawl_path = exports_dir / crawl_file.name
            crawl_file.seek(0)
            crawl_path.write_bytes(crawl_file.read())
            logs.append(f'<span class="log-ok">✓</span> Crawl export saved → {crawl_file.name}')

        # ── STEP 1: merge_original_exports ────────────────────────────────
        logs.append('<span class="log-info">▸ STEP 1 — Merging raw export files…</span>')
        try:
            sys.path.insert(0, str(base))
            import importlib, types

            def run_merge(date_str_: str, base_dir_: Path) -> str:
                folder = base_dir_ / date_str_ / "exports"
                files = sorted(
                    f for f in os.listdir(folder)
                    if f.lower().endswith((".xlsx", ".xls"))
                    and not f.startswith("~$")
                    and not f.lower().startswith("merged_exports")
                    and not f.lower().startswith("merged_original_export")
                )
                if not files:
                    raise FileNotFoundError("No export files found in uploads.")
                frames = [pd.read_excel(folder / f) for f in files]
                merged = pd.concat(frames, ignore_index=True)
                out = folder / f"merged_original_export_{date_str_}.xlsx"
                merged.to_excel(out, index=False)
                return str(out)

            merged_orig_path = run_merge(date_str, base)
            logs.append(f'<span class="log-ok">✓</span> merged_original_export_{date_str}.xlsx created ({pd.read_excel(merged_orig_path).shape[0]} rows)')
        except Exception as e:
            logs.append(f'<span class="log-err">✗ STEP 1 FAILED: {e}</span>')
            return None, logs

        # ── STEP 2: reshape_merged_for_kpi_order ──────────────────────────
        logs.append('<span class="log-info">▸ STEP 2 — Reshaping for KPI order…</span>')
        try:
            KPI_ORDER = [
                "New Listings",
                "Used Listings",
                "New Cars Listings ",
                "Used Cars Listings ",
                "New Motorbikes Listings ",
                "Used Motorbikes Listings ",
                "New Trucks Listings ",
                "Used Trucks Listings ",
                "New Motorhomes Listings ",
                "Used Motorhomes Listings ",
            ]
            kpi_map = {
                "new": "New Listings",
                "used": "Used Listings",
                "new cars": "New Cars Listings ",
                "used cars": "Used Cars Listings ",
                "new motorbikes": "New Motorbikes Listings ",
                "used motorbikes": "Used Motorbikes Listings ",
                "new truck": "New Trucks Listings ",
                "used truck": "Used Trucks Listings ",
                "new trucks": "New Trucks Listings ",
                "used trucks": "Used Trucks Listings ",
                "new caravan": "New Motorhomes Listings ",
                "used caravan": "Used Motorhomes Listings ",
                "new motorhomes": "New Motorhomes Listings ",
                "used motorhomes": "Used Motorhomes Listings ",
            }

            df = pd.read_excel(merged_orig_path)
            required = {"Website", "Date Pige", "Name Of Detail Pige", "Detail Pige"}
            missing_cols = required - set(df.columns)
            if missing_cols:
                raise ValueError(f"Missing columns: {missing_cols}")

            twb = pd.read_excel(tpl_dest, sheet_name=0, engine="openpyxl")
            if "Dealer" not in twb.columns:
                raise ValueError("Template missing 'Dealer' column.")
            dealer_order = twb["Dealer"].dropna().astype(str).str.strip().drop_duplicates().tolist()

            raw = df["Name Of Detail Pige"].astype(str).str.strip().str.lower()
            df["KPI_std"] = raw.map(kpi_map)
            df = df[df["KPI_std"].isin(KPI_ORDER)].copy()
            df["Website"] = df["Website"].astype(str).str.strip()

            cur_map = (
                df[["Website", "KPI_std", "Detail Pige"]]
                .dropna(subset=["Detail Pige", "Website", "KPI_std"])
                .drop_duplicates(subset=["Website", "KPI_std"])
                .set_index(["Website", "KPI_std"])["Detail Pige"]
                .to_dict()
            )
            date_map = (
                df[["Website", "Date Pige"]]
                .dropna(subset=["Website"])
                .drop_duplicates(subset=["Website"])
                .set_index("Website")["Date Pige"]
                .to_dict()
            )

            rows = []
            for dealer in dealer_order:
                d = str(dealer).strip()
                dv = date_map.get(d)
                for kpi in KPI_ORDER:
                    val = cur_map.get((d, kpi), "N/A")
                    if pd.isna(val):
                        val = "N/A"
                    rows.append({"Website": d, "Date Pige": dv, "Name Of Detail Pige": kpi, "Detail Pige": val})

            result = pd.DataFrame(rows, columns=["Website", "Date Pige", "Name Of Detail Pige", "Detail Pige"])
            out2 = exports_dir / f"merged_exports_{date_str}.xlsx"
            result.to_excel(out2, index=False)
            logs.append(f'<span class="log-ok">✓</span> merged_exports_{date_str}.xlsx created ({len(result)} rows, {len(dealer_order)} dealers)')
        except Exception as e:
            logs.append(f'<span class="log-err">✗ STEP 2 FAILED: {e}</span>')
            return None, logs

        # ── STEP 3: build_from_exports_generic ────────────────────────────
        logs.append('<span class="log-info">▸ STEP 3 — Building final Template…</span>')
        try:
            import openpyxl
            from collections import defaultdict, OrderedDict

            merged_df = pd.read_excel(out2)

            wb = openpyxl.load_workbook(tpl_dest, data_only=False)
            ws = wb[wb.sheetnames[0]]

            # Previous week values from template
            prev_wb = openpyxl.load_workbook(tpl_dest, data_only=True)
            prev_ws = prev_wb[prev_wb.sheetnames[0]]

            DEALER_COL, KPI_COL, VALUE_COL, LAST_WEEK_COL = 5, 9, 10, 11
            YEAR_COL, MONTH_COL, DAY_COL = 1, 2, 3

            prev_value_lookup = {}
            for prow in prev_ws.iter_rows(min_row=2, max_row=prev_ws.max_row, values_only=True):
                if not prow:
                    continue
                pd_ = str(prow[DEALER_COL-1]).strip() if prow[DEALER_COL-1] else ""
                pk_ = str(prow[KPI_COL-1]).strip() if prow[KPI_COL-1] else ""
                if pd_ and pk_:
                    prev_value_lookup[(pd_, pk_)] = prow[VALUE_COL-1]

            # Build agg lookup from merged_df
            if "Status" in merged_df.columns:
                merged_df = merged_df[merged_df["Status"].astype(str).str.lower() == "completed"]
            merged_df = merged_df[merged_df["Detail Pige"].notna()]

            def norm(s):
                return (s or "").strip().lower()

            agg = defaultdict(float)
            for _, row in merged_df.iterrows():
                dealer_ = str(row["Website"]).strip()
                raw_name = str(row["Name Of Detail Pige"]).strip()
                kpi_base = norm(raw_name.replace("Listings", "").replace("listings", ""))
                try:
                    val_ = float(row["Detail Pige"])
                except Exception:
                    continue
                agg[(dealer_, kpi_base)] += val_

            dt = datetime.strptime(date_str, "%d.%m.%Y")
            year, month, day = dt.year, dt.month, dt.day

            # Load base rows from template
            template_rows = []
            for row in prev_ws.iter_rows(min_row=2, max_row=prev_ws.max_row, values_only=True):
                if all(v is None for v in row):
                    continue
                template_rows.append(list(row))

            ws.cell(row=1, column=LAST_WEEK_COL, value="Last week Value")

            for row_idx, row in enumerate(template_rows, start=2):
                dealer_str = str(row[DEALER_COL-1]).strip() if row[DEALER_COL-1] else ""
                kpi_str = str(row[KPI_COL-1]).strip() if row[KPI_COL-1] else ""
                prev_val = prev_value_lookup.get((dealer_str, kpi_str), "N/A")

                for col_idx in range(1, VALUE_COL + 1):
                    cv = row[col_idx-1] if len(row) >= col_idx else None
                    ws.cell(row=row_idx, column=col_idx, value=cv)

                ws.cell(row=row_idx, column=YEAR_COL, value=year)
                ws.cell(row=row_idx, column=MONTH_COL, value=month)
                ws.cell(row=row_idx, column=DAY_COL, value=day)
                ws.cell(row=row_idx, column=LAST_WEEK_COL, value=prev_val)

                if not dealer_str or not kpi_str:
                    ws.cell(row=row_idx, column=VALUE_COL, value="N/A")
                    continue

                kpi_base = norm(kpi_str.replace("Listings", "").replace("listings", ""))
                val_ = agg.get((dealer_str, kpi_base))
                ws.cell(row=row_idx, column=VALUE_COL, value=val_ if val_ is not None else "N/A")

            exp_last = len(template_rows) + 1
            if ws.max_row > exp_last:
                ws.delete_rows(exp_last + 1, ws.max_row - exp_last)

            out3 = deliver_dir / f"Template_{date_str}.xlsx"
            wb.save(out3)
            logs.append(f'<span class="log-ok">✓</span> Template_{date_str}.xlsx generated ({len(template_rows)} rows)')
        except Exception as e:
            logs.append(f'<span class="log-err">✗ STEP 3 FAILED: {e}</span>')
            return None, logs

        # ── STEP 4 (optional): crawl update ───────────────────────────────
        if crawl_path and crawl_path.exists():
            logs.append('<span class="log-info">▸ STEP 4 — Applying crawl export update…</span>')
            try:
                import openpyxl

                crawl_wb2 = openpyxl.load_workbook(crawl_path, data_only=True)
                crawl_ws2 = crawl_wb2[crawl_wb2.sheetnames[0]]
                crawl_lookup = {}
                for crow in crawl_ws2.iter_rows(min_row=2, values_only=True):
                    if not crow or len(crow) < 10:
                        continue
                    cd = norm(str(crow[4] or ""))
                    ck = norm(str(crow[8] or ""))
                    cv = crow[9]
                    if cd and ck and cv is not None:
                        crawl_lookup[(cd, ck)] = cv

                wb4 = openpyxl.load_workbook(out3, data_only=False)
                ws4 = wb4[wb4.sheetnames[0]]
                updated = 0
                for ri in range(2, ws4.max_row + 1):
                    d_ = norm(str(ws4.cell(ri, DEALER_COL).value or ""))
                    k_ = norm(str(ws4.cell(ri, KPI_COL).value or ""))
                    if (d_, k_) in crawl_lookup:
                        ws4.cell(ri, VALUE_COL, value=crawl_lookup[(d_, k_)])
                        updated += 1
                wb4.save(out3)
                logs.append(f'<span class="log-ok">✓</span> Crawl update applied — {updated} cells overwritten')
            except Exception as e:
                logs.append(f'<span class="log-err">⚠ STEP 4 WARNING: {e}</span>')

        # Read final file bytes
        output_bytes = out3.read_bytes()
        logs.append(f'<span class="log-ok">✓ DONE — Template_{date_str}.xlsx ready for download</span>')
        return output_bytes, logs


# ── UI ────────────────────────────────────────────────────────────────────────
st.markdown('<div class="page-title">Run <span class="accent">Pipeline</span></div>', unsafe_allow_html=True)
st.markdown('<div class="page-sub">Upload exports + template → get your weekly deliverable</div>', unsafe_allow_html=True)

col_form, col_log = st.columns([1, 1], gap="large")

with col_form:
    st.markdown("#### Configuration")

    date_input = st.text_input(
        "Week Date (DD.MM.YYYY)",
        placeholder="e.g. 28.04.2026",
        help="The date of the current week's exports",
    )

    st.markdown('<div class="upload-label">Detail Export Files *</div>', unsafe_allow_html=True)
    export_uploads = st.file_uploader(
        "Upload one or more detail export files",
        type=["xlsx", "xls"],
        accept_multiple_files=True,
        key="exports",
        label_visibility="collapsed",
    )
    st.markdown('<div class="upload-hint">One or more <code>detail_*.xlsx</code> files from the exports folder</div>', unsafe_allow_html=True)

    st.markdown('<div class="upload-label" style="margin-top:1rem;">Previous Week Template *</div>', unsafe_allow_html=True)
    template_upload = st.file_uploader(
        "Upload the previous week Template file",
        type=["xlsx"],
        key="template",
        label_visibility="collapsed",
    )
    st.markdown('<div class="upload-hint">Last week\'s <code>Template_*.xlsx</code> — its Values become Last week Values</div>', unsafe_allow_html=True)

    st.markdown('<div class="upload-label" style="margin-top:1rem;">Crawl Export (Optional)</div>', unsafe_allow_html=True)
    crawl_upload = st.file_uploader(
        "Upload crawl export file",
        type=["xlsx"],
        key="crawl",
        label_visibility="collapsed",
    )
    st.markdown('<div class="upload-hint">Optional <code>crawl_export_*.xlsx</code> to overwrite Values after pipeline</div>', unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    run_disabled = not (date_input and export_uploads and template_upload)
    run_btn = st.button("▶  Run Pipeline", disabled=run_disabled, use_container_width=True)

with col_log:
    st.markdown("#### Pipeline Log")

    log_placeholder = st.empty()
    result_placeholder = st.empty()

    with log_placeholder.container():
        st.markdown(
            '<div class="log-box"><span class="log-info">Waiting for pipeline to start…</span></div>',
            unsafe_allow_html=True,
        )

if run_btn:
    # Validate date
    try:
        datetime.strptime(date_input.strip(), "%d.%m.%Y")
    except ValueError:
        st.error("Invalid date format. Use DD.MM.YYYY (e.g. 28.04.2026)")
        st.stop()

    with st.spinner("Running pipeline…"):
        output_bytes, logs = run_pipeline(
            date_str=date_input.strip(),
            export_files=export_uploads,
            template_file=template_upload,
            crawl_file=crawl_upload,
        )

    log_html = "<br>".join(logs)
    with log_placeholder.container():
        st.markdown(f'<div class="log-box">{log_html}</div>', unsafe_allow_html=True)

    if output_bytes:
        with result_placeholder.container():
            st.markdown(
                f'<div class="result-card">'
                f'<div class="result-title">✓ Output Ready</div>'
                f'Template_{date_input.strip()}.xlsx is ready for download.'
                f'</div>',
                unsafe_allow_html=True,
            )
            st.download_button(
                label=f"⬇  Download Template_{date_input.strip()}.xlsx",
                data=output_bytes,
                file_name=f"Template_{date_input.strip()}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
    else:
        st.error("Pipeline failed. Check the log above for details.")
