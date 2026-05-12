import os
import sys
from collections import OrderedDict, defaultdict
from datetime import datetime
from typing import Dict, Tuple

import openpyxl
import pandas as pd


BASE_DIR = os.path.dirname(__file__)
TEMPLATE_PATH = os.path.join(BASE_DIR, "Template_file.xlsx")


def norm(s: str) -> str:
    return (s or "").strip().lower()


def kpi_from_detail_name(name: str, template_kpis_norm_map: Dict[str, str]) -> str | None:
    """
    Export'taki Name Of Detail Pige -> Template'teki KPI string.
    Mantık: sadece 'listings' kelimesi eksik; kelimeleri title-case yapıp sonuna ' Listings' ekliyoruz,
    sonra Template KPI'lar içinde normalize eşleşme arıyoruz.
    """
    base = (name or "").strip()
    if not base:
        return None
    # Örn: "new cars" -> "New Cars Listings"
    candidate = base.title() + " Listings"
    cand_norm = norm(candidate)
    return template_kpis_norm_map.get(cand_norm)


def merge_exports_for_date(date_str: str) -> pd.DataFrame:
    """
    Artık export tekil dosyalarını burada birleştirmiyoruz.
    Kullanıcı, ilgili tarih için `merged_exports_<date>.xlsx` dosyasını
    önceden hazırlıyor (örneğin `reshape_merged_for_kpi_order.py` ile).

    Bu fonksiyon, sadece o hazır merged export dosyasını OKUR ve
    Template güncellemesi için DataFrame döner.
    """
    folder = os.path.join(BASE_DIR, date_str, "exports")
    merged_path = os.path.join(folder, f"merged_exports_{date_str}.xlsx")
    if not os.path.isfile(merged_path):
        raise FileNotFoundError(
            f"Merged export file not found: {merged_path}. "
            "Önce merged export dosyasını oluşturup kaydetmelisiniz."
        )

    return pd.read_excel(merged_path)


def resolve_template_path(date_str: str) -> str:
    """
    Hedef tarihten onceki en guncel deliver/Template_<date>.xlsx
    dosyasini bulur ve onu template baz olarak dondurur.
    Bulamazsa varsayilan Template_file.xlsx kullanilir.
    """
    target_dt = datetime.strptime(date_str, "%d.%m.%Y")
    fallback = TEMPLATE_PATH

    best_candidate: tuple[datetime, str] | None = None
    for entry in os.listdir(BASE_DIR):
        folder_path = os.path.join(BASE_DIR, entry)
        if not os.path.isdir(folder_path):
            continue
        try:
            folder_dt = datetime.strptime(entry, "%d.%m.%Y")
        except ValueError:
            continue
        if folder_dt >= target_dt:
            continue

        candidate = os.path.join(folder_path, "deliver", f"Template_{entry}.xlsx")
        if not os.path.isfile(candidate):
            continue

        if best_candidate is None or folder_dt > best_candidate[0]:
            best_candidate = (folder_dt, candidate)

    if best_candidate is not None:
        return best_candidate[1]
    return fallback


def load_base_rows_for_target_week(date_str: str, fallback_template_path: str) -> list[list]:
    """
    Template satir setini secilen base template dosyasindan alir.
    (Tum website satirlarini korumak icin baz hep full template olur.)
    """
    wb = openpyxl.load_workbook(fallback_template_path, data_only=False)
    ws = wb[wb.sheetnames[0]]

    rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if all(v is None for v in row):
            continue
        rows.append(list(row))
    return rows


def build_lookup_from_merged(df: pd.DataFrame) -> Tuple[OrderedDict, Dict[Tuple[str, str], float]]:
    """
    - Dealer = Website (birebir)
    - KPI eşleşmesi:
        * Export tarafı: Name Of Detail Pige
        * Template tarafı: KPI
        * Her ikisinde de 'listings' kelimesi ve boşluklar normalize edilerek karşılaştırılır.
          Yani sadece gerçek kategori metni eşleştirilir (ör. 'new cars').
    - Value = Detail Pige (numeric, satırlar toplanır)
    Dealer + KPI-basesi bazında toplam değer sözlüğü döndürür.
    """
    required_cols = {"Website", "Name Of Detail Pige", "Detail Pige", "Date Pige"}
    missing = required_cols - set(df.columns)
    if missing:
        raise ValueError(f"Missing columns in exports: {missing}")

    # Sadece Completed + numeric Detail Pige satırları al
    df = df.copy()
    if "Status" in df.columns:
        df = df[df["Status"].astype(str).str.lower() == "completed"]
    df = df[df["Detail Pige"].notna()]

    # Dealer sırası (Website kolonundan - orijinal isimleri koruyoruz)
    dealers_ordered = list(
        OrderedDict.fromkeys(df["Website"].astype(str).str.strip().tolist())
    )

    # KPI base: 'listings' kelimesini ve boşluklarını atarak normalize et
    agg: Dict[Tuple[str, str], float] = defaultdict(float)
    for _, row in df.iterrows():
        dealer = str(row["Website"]).strip()
        raw_name = str(row["Name Of Detail Pige"]).strip()
        if not dealer or not raw_name:
            continue
        # Örn: 'New Cars Listings' -> 'new cars'
        kpi_base = norm(
            raw_name.replace("Listings", "").replace("listings", "")
        )
        try:
            val = float(row["Detail Pige"])
        except Exception:
            continue
        agg[(dealer, kpi_base)] += val

    return OrderedDict((d, None) for d in dealers_ordered), agg


def build_template_for_date(date_str: str) -> None:
    """
    - exports klasöründeki dosyaları merge eder
    - Template_file.xlsx yapısına göre yeni bir dosya üretir
    - Dealer sırasını Template'teki sıraya göre kullanır
    - KPI eşleşmesi: Template.KPI ile exporttan türetilen KPI birebir eşleşecek
    - Değer yoksa 'N/A' yazar
    - Year / Month / Day kolonlarını verilen tarihe göre günceller
    - Çıktıyı <date>\\deliver içine kaydeder.
    """
    # 1) Exports'tan lookup oluştur
    merged_df = merge_exports_for_date(date_str)

    # Template'i oku (onceki en guncel hafta template'i tercih edilir)
    selected_template_path = resolve_template_path(date_str)
    if not os.path.isfile(selected_template_path):
        raise FileNotFoundError(f"Template file not found: {selected_template_path}")

    wb = openpyxl.load_workbook(selected_template_path, data_only=False)
    ws = wb[wb.sheetnames[0]]

    _, agg = build_lookup_from_merged(merged_df)

    # Bu dosya haftalık template üretimi için kullanılıyor.
    # Kullanıcı verdiği `date_str` değerindeki yıl/ay/günü aynen uygular (ör. 24.03.2026 -> day=24).
    dt = datetime.strptime(date_str, "%d.%m.%Y")
    year, month, day = dt.year, dt.month, dt.day

    # 2) Template'teki dealer ve KPI sırasını oku
    # Bu hafta crawl export varsa website/dealer satir setini oradan al
    template_rows = load_base_rows_for_target_week(date_str, selected_template_path)

    # Template sütun indeksleri (1-based)
    # ['Year', 'Month', 'Day', 'customer_number', 'Dealer', 'Type', 'Category', 'Vehicle_condition', 'KPI', 'Value']
    YEAR_COL = 1
    MONTH_COL = 2
    DAY_COL = 3
    DEALER_COL = 5
    KPI_COL = 9
    VALUE_COL = 10
    LAST_WEEK_COL = 11

    # Başlığa "Last week Value" kolonu ekle
    ws.cell(row=1, column=LAST_WEEK_COL, value="Last week Value")

    # Onceki hafta template'indeki Value degerlerini (last week) lookup'a al
    prev_wb = openpyxl.load_workbook(selected_template_path, data_only=True)
    prev_ws = prev_wb[prev_wb.sheetnames[0]]
    prev_value_lookup: Dict[Tuple[str, str], float | str] = {}
    for prow in prev_ws.iter_rows(min_row=2, max_row=prev_ws.max_row, values_only=True):
        if not prow:
            continue
        pdealer = str(prow[DEALER_COL - 1]).strip() if prow[DEALER_COL - 1] is not None else ""
        pkpi = str(prow[KPI_COL - 1]).strip() if prow[KPI_COL - 1] is not None else ""
        if not pdealer or not pkpi:
            continue
        prev_value_lookup[(pdealer, pkpi)] = prow[VALUE_COL - 1]

    # Dealer sırasını template'e göre kullan
    # Aynı zamanda customer_number, Type, Category, Vehicle_condition gibi alanları template'ten birebir koruyoruz.
    for row_idx, row in enumerate(template_rows, start=2):
        dealer = row[DEALER_COL - 1]
        kpi = row[KPI_COL - 1]
        dealer_str = str(dealer).strip() if dealer is not None else ""
        kpi_str = str(kpi).strip() if kpi is not None else ""
        previous_value = prev_value_lookup.get((dealer_str, kpi_str), "N/A")

        # Bu haftanin baz satirini (ilk 10 kolonu) aynen yaz
        for col_idx in range(1, VALUE_COL + 1):
            cell_value = row[col_idx - 1] if len(row) >= col_idx else None
            ws.cell(row=row_idx, column=col_idx, value=cell_value)

        # Tarihi güncelle
        ws.cell(row=row_idx, column=YEAR_COL, value=year)
        ws.cell(row=row_idx, column=MONTH_COL, value=month)
        ws.cell(row=row_idx, column=DAY_COL, value=day)

        # Eski değeri Last week Value kolonuna yaz
        ws.cell(row=row_idx, column=LAST_WEEK_COL, value=previous_value)

        if not dealer_str or not kpi_str:
            ws.cell(row=row_idx, column=VALUE_COL, value="N/A")
            continue

        # Lookup key: Dealer (birebir) + KPI base (listings'siz, normalize)
        kpi_base = norm(
            kpi_str.replace("Listings", "").replace("listings", "")
        )

        # Lookup'ta varsa değeri yaz, yoksa N/A
        val = agg.get((dealer_str, kpi_base))
        if val is None:
            ws.cell(row=row_idx, column=VALUE_COL, value="N/A")
        else:
            ws.cell(row=row_idx, column=VALUE_COL, value=val)

    # Baz satir sayisindan fazla kalan eski satirlari temizle
    expected_last_row = len(template_rows) + 1
    if ws.max_row > expected_last_row:
        ws.delete_rows(expected_last_row + 1, ws.max_row - expected_last_row)

    # 3) Yeni dosyayı deliver klasörüne kaydet
    deliver_dir = os.path.join(BASE_DIR, date_str, "deliver")
    os.makedirs(deliver_dir, exist_ok=True)
    out_path = os.path.join(deliver_dir, f"Template_{date_str}.xlsx")
    wb.save(out_path)
    print(f"Template base used: {selected_template_path}")
    print(f"Template file for {date_str} created at: {out_path}")


if __name__ == "__main__":
    # Örn: python build_from_exports_generic.py 24.03.2026
    date = sys.argv[1] if len(sys.argv) > 1 else "17.03.2026"
    build_template_for_date(date)

