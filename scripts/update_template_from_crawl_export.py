import argparse
from pathlib import Path

import openpyxl


def norm_text(value) -> str:
    return str(value or "").strip().lower()


def has_meaningful_value(value) -> bool:
    if value is None:
        return False
    if isinstance(value, str) and norm_text(value) == "":
        return False
    return True


def update_template_from_crawl(crawl_path: Path, template_path: Path, output_path: Path) -> None:
    crawl_wb = openpyxl.load_workbook(crawl_path, data_only=True)
    crawl_ws = crawl_wb[crawl_wb.sheetnames[0]]

    template_wb = openpyxl.load_workbook(template_path, data_only=False)
    template_ws = template_wb[template_wb.sheetnames[0]]

    # Crawl columns expected: Dealer, KPI, Value
    crawl_lookup = {}
    for row in crawl_ws.iter_rows(min_row=2, max_row=crawl_ws.max_row, values_only=True):
        if row is None or len(row) < 10:
            continue
        dealer = norm_text(row[4])  # Dealer
        kpi = norm_text(row[8])     # KPI
        value = row[9]              # Value
        if not dealer or not kpi or not has_meaningful_value(value):
            continue
        crawl_lookup[(dealer, kpi)] = value

    updated_rows = 0
    missing_rows = 0

    # Template columns expected: Dealer (5), KPI (9), Value (10)
    for row_idx in range(2, template_ws.max_row + 1):
        dealer = norm_text(template_ws.cell(row=row_idx, column=5).value)
        kpi = norm_text(template_ws.cell(row=row_idx, column=9).value)
        if not dealer or not kpi:
            continue

        key = (dealer, kpi)
        if key in crawl_lookup:
            template_ws.cell(row=row_idx, column=10, value=crawl_lookup[key])
            updated_rows += 1
        else:
            missing_rows += 1

    output_path.parent.mkdir(parents=True, exist_ok=True)
    template_wb.save(output_path)

    print(f"Crawl rows mapped: {len(crawl_lookup)}")
    print(f"Template rows updated: {updated_rows}")
    print(f"Template rows without crawl match: {missing_rows}")
    print(f"Saved: {output_path}")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Update template Value column using crawl export by Dealer + KPI."
    )
    parser.add_argument("--crawl", required=True, help="Path to crawl_export_*.xlsx")
    parser.add_argument("--template", required=True, help="Path to Template_*.xlsx")
    parser.add_argument(
        "--output",
        required=False,
        help="Output path. If omitted, overwrites template file.",
    )
    args = parser.parse_args()

    crawl_path = Path(args.crawl)
    template_path = Path(args.template)
    output_path = Path(args.output) if args.output else template_path

    if not crawl_path.is_file():
        raise FileNotFoundError(f"Crawl file not found: {crawl_path}")
    if not template_path.is_file():
        raise FileNotFoundError(f"Template file not found: {template_path}")

    update_template_from_crawl(crawl_path, template_path, output_path)


if __name__ == "__main__":
    main()
